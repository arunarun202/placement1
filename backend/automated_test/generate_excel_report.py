"""
Convert DAST report.json → DAST_Security_Report.xlsx
Reads report.json from the same directory and writes a formatted Excel file.
"""
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(__file__)
REPORT_JSON = os.path.join(BASE_DIR, "report.json")
OUTPUT_XLSX = os.path.join(BASE_DIR, "DAST_Security_Report_v2.xlsx")

# ── Load data ──────────────────────────────────────────────
with open(REPORT_JSON, encoding="utf-8") as f:
    records = json.load(f)

# ── Colour palette ─────────────────────────────────────────
COLOURS = {
    "header_bg":    "1F3864",   # dark navy
    "header_fg":    "FFFFFF",
    "critical_bg":  "C00000",   # dark red
    "critical_fg":  "FFFFFF",
    "high_bg":      "FF4444",
    "high_fg":      "FFFFFF",
    "medium_bg":    "FF8C00",
    "medium_fg":    "FFFFFF",
    "low_bg":       "FFC000",
    "low_fg":       "000000",
    "pass_bg":      "E2EFDA",   # light green
    "fail_bg":      "FCE4D6",   # light red
    "alt_row":      "F2F2F2",
    "title_bg":     "2E75B6",
    "title_fg":     "FFFFFF",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "Executive Summary"

ws_sum.column_dimensions["A"].width = 30
ws_sum.column_dimensions["B"].width = 18

# Title
ws_sum.merge_cells("A1:B1")
c = ws_sum["A1"]
c.value = "DAST Security Report — Placement Predictor API"
c.fill = fill(COLOURS["title_bg"])
c.font = font(bold=True, color=COLOURS["title_fg"], size=14)
c.alignment = center()
ws_sum.row_dimensions[1].height = 32

ws_sum.merge_cells("A2:B2")
c = ws_sum["A2"]
c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
c.font = font(color="555555", size=10)
c.alignment = center()
ws_sum.row_dimensions[2].height = 18

ws_sum.append([])  # blank row

# Stats
total      = len(records)
findings   = [r for r in records if r.get("finding")]
critical_n = sum(1 for r in findings if r.get("severity") == "CRITICAL")
high_n     = sum(1 for r in findings if r.get("severity") == "HIGH")
medium_n   = sum(1 for r in findings if r.get("severity") == "MEDIUM")
low_n      = sum(1 for r in findings if r.get("severity") == "LOW")
pass_n     = total - len(findings)

stats = [
    ("Total Tests / Findings",          f"{total} tests  |  {len(findings)} findings"),
    ("🔴 CRITICAL Findings",            critical_n),
    ("🔴 HIGH Findings",                high_n),
    ("🟠 MEDIUM Findings",              medium_n),
    ("🟡 LOW Findings",                 low_n),
    ("✓  Passed (no finding)",          pass_n),
]

sev_colours = {
    "🔴 CRITICAL Findings": (COLOURS["critical_bg"], COLOURS["critical_fg"]),
    "🔴 HIGH Findings":     (COLOURS["high_bg"],     COLOURS["high_fg"]),
    "🟠 MEDIUM Findings":   (COLOURS["medium_bg"],   COLOURS["medium_fg"]),
    "🟡 LOW Findings":      (COLOURS["low_bg"],      COLOURS["low_fg"]),
}

for label, value in stats:
    row_idx = ws_sum.max_row + 1
    ws_sum.append([label, value])
    lc = ws_sum.cell(row=row_idx, column=1)
    vc = ws_sum.cell(row=row_idx, column=2)
    lc.font = font(bold=True)
    lc.alignment = left()
    lc.border = border
    vc.alignment = center()
    vc.border = border
    if label in sev_colours:
        bg, fg = sev_colours[label]
        lc.fill = fill(bg); lc.font = font(bold=True, color=fg)
        vc.fill = fill(bg); vc.font = font(bold=True, color=fg)
    ws_sum.row_dimensions[row_idx].height = 20

# Top issues section
ws_sum.append([])
row_idx = ws_sum.max_row + 1
ws_sum.merge_cells(f"A{row_idx}:B{row_idx}")
c = ws_sum.cell(row=row_idx, column=1)
c.value = "TOP ISSUES TO FIX FIRST"
c.fill = fill(COLOURS["header_bg"])
c.font = font(bold=True, color=COLOURS["header_fg"])
c.alignment = center()
ws_sum.row_dimensions[row_idx].height = 22

priority_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]
findings_sorted = sorted(findings, key=lambda r: priority_order.index(r.get("severity","LOW")) if r.get("severity") in priority_order else 99)
for i, r in enumerate(findings_sorted[:10], 1):
    row_idx = ws_sum.max_row + 1
    sev = r.get("severity", "")
    ws_sum.append([f"{i}. [{sev}] {r.get('endpoint','')}", r.get("note", "")[:120]])
    lc = ws_sum.cell(row=row_idx, column=1)
    nc = ws_sum.cell(row=row_idx, column=2)
    bg = COLOURS.get(f"{sev.lower()}_bg", "FFFFFF")
    fg = COLOURS.get(f"{sev.lower()}_fg", "000000")
    lc.fill = fill(bg); lc.font = font(bold=True, color=fg, size=10)
    nc.fill = fill(COLOURS["fail_bg"]); nc.font = font(size=10)
    lc.alignment = left(); nc.alignment = left()
    lc.border = border; nc.border = border
    ws_sum.row_dimensions[row_idx].height = 28

# ═══════════════════════════════════════════════════════════
# SHEET 2 — ALL TEST RESULTS
# ═══════════════════════════════════════════════════════════
ws = wb.create_sheet("All Test Results")

COLS = ["ID", "Endpoint", "Method", "Role", "Category",
        "Status Code", "Expected Status", "Finding?", "Severity",
        "Response Time (ms)", "Note", "Timestamp"]

COL_WIDTHS = [5, 32, 10, 12, 18, 13, 15, 10, 12, 20, 60, 22]

for i, (col, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Header row
for col_idx, col_name in enumerate(COLS, 1):
    c = ws.cell(row=1, column=col_idx, value=col_name)
    c.fill = fill(COLOURS["header_bg"])
    c.font = font(bold=True, color=COLOURS["header_fg"])
    c.alignment = center()
    c.border = border
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# Data rows
for row_num, rec in enumerate(records, 2):
    is_finding = rec.get("finding", False)
    sev        = rec.get("severity", "none")
    row_bg     = COLOURS["fail_bg"] if is_finding else (COLOURS["alt_row"] if row_num % 2 == 0 else "FFFFFF")

    values = [
        row_num - 1,
        rec.get("endpoint", ""),
        rec.get("method", ""),
        rec.get("role", ""),
        rec.get("test_category", ""),
        str(rec.get("status", "")),
        str(rec.get("expected_status", "")),
        "✗ YES" if is_finding else "✓ No",
        sev,
        rec.get("response_time_ms", 0),
        rec.get("note", ""),
        rec.get("timestamp", ""),
    ]

    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.alignment = left() if col_idx in (2, 11) else center()
        c.border = border
        c.fill = fill(row_bg)
        c.font = font(size=10)

    # Colour the Severity and Finding columns
    sev_cell = ws.cell(row=row_num, column=9)
    fin_cell = ws.cell(row=row_num, column=8)
    if sev == "CRITICAL":
        sev_cell.fill = fill(COLOURS["critical_bg"]); sev_cell.font = font(bold=True, color=COLOURS["critical_fg"], size=10)
        fin_cell.fill = fill(COLOURS["critical_bg"]); fin_cell.font = font(bold=True, color=COLOURS["critical_fg"], size=10)
    elif sev == "HIGH":
        sev_cell.fill = fill(COLOURS["high_bg"]); sev_cell.font = font(bold=True, color=COLOURS["high_fg"], size=10)
        fin_cell.fill = fill(COLOURS["high_bg"]); fin_cell.font = font(bold=True, color=COLOURS["high_fg"], size=10)
    elif sev == "MEDIUM":
        sev_cell.fill = fill(COLOURS["medium_bg"]); sev_cell.font = font(bold=True, color=COLOURS["medium_fg"], size=10)
        fin_cell.fill = fill(COLOURS["medium_bg"]); fin_cell.font = font(bold=True, color=COLOURS["medium_fg"], size=10)

    ws.row_dimensions[row_num].height = 36

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ═══════════════════════════════════════════════════════════
# SHEET 3 — FINDINGS ONLY
# ═══════════════════════════════════════════════════════════
ws_f = wb.create_sheet("Findings Only")
for i, (col, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
    ws_f.column_dimensions[get_column_letter(i)].width = w

for col_idx, col_name in enumerate(COLS, 1):
    c = ws_f.cell(row=1, column=col_idx, value=col_name)
    c.fill = fill(COLOURS["critical_bg"])
    c.font = font(bold=True, color="FFFFFF")
    c.alignment = center()
    c.border = border
ws_f.row_dimensions[1].height = 28
ws_f.freeze_panes = "A2"

for row_num, rec in enumerate(findings_sorted, 2):
    sev = rec.get("severity", "none")
    if sev == "CRITICAL":   bg = COLOURS["critical_bg"]; fg = COLOURS["critical_fg"]
    elif sev == "HIGH":     bg = COLOURS["high_bg"];     fg = COLOURS["high_fg"]
    elif sev == "MEDIUM":   bg = COLOURS["medium_bg"];   fg = COLOURS["medium_fg"]
    else:                   bg = COLOURS["low_bg"];      fg = COLOURS["low_fg"]

    values = [
        row_num - 1,
        rec.get("endpoint", ""),
        rec.get("method", ""),
        rec.get("role", ""),
        rec.get("test_category", ""),
        str(rec.get("status", "")),
        str(rec.get("expected_status", "")),
        "✗ FINDING",
        sev,
        rec.get("response_time_ms", 0),
        rec.get("note", ""),
        rec.get("timestamp", ""),
    ]
    for col_idx, val in enumerate(values, 1):
        c = ws_f.cell(row=row_num, column=col_idx, value=val)
        c.fill = fill(bg)
        c.font = font(bold=(col_idx == 9), color=fg, size=10)
        c.alignment = left() if col_idx in (2, 11) else center()
        c.border = border
    ws_f.row_dimensions[row_num].height = 40

ws_f.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ═══════════════════════════════════════════════════════════
# SHEET 4 — ENDPOINT INVENTORY
# ═══════════════════════════════════════════════════════════
ws_ep = wb.create_sheet("Endpoint Inventory")
ws_ep.column_dimensions["A"].width = 5
ws_ep.column_dimensions["B"].width = 12
ws_ep.column_dimensions["C"].width = 35
ws_ep.column_dimensions["D"].width = 20

endpoints = [
    ("#", "Method", "Path", "Auth Required?"),
    (1, "GET",    "/",                "Public"),
    (2, "POST",   "/auth/login",      "Public"),
    (3, "POST",   "/auth/register",   "Public"),
    (4, "GET",    "/auth/me",         "✅ JWT required"),
    (5, "POST",   "/predict/",        "✅ JWT required"),
    (6, "GET",    "/predict/",        "✅ JWT required"),
    (7, "POST",   "/resume/upload",   "✅ JWT required"),
    (8, "GET",    "/resume/",         "✅ JWT required"),
    (9, "GET",    "/resume/{id}",     "✅ JWT required"),
    (10,"DELETE", "/resume/{id}",     "✅ JWT required"),
    (11,"POST",   "/chat/",           "✅ JWT required"),
    (12,"GET",    "/chat/",           "✅ JWT required ⚠ IDOR"),
    (13,"PUT",    "/profile/",        "✅ JWT required"),
]
for row_num, row in enumerate(endpoints, 1):
    for col_idx, val in enumerate(row, 1):
        c = ws_ep.cell(row=row_num, column=col_idx, value=val)
        c.border = border
        c.alignment = center()
        if row_num == 1:
            c.fill = fill(COLOURS["header_bg"]); c.font = font(bold=True, color="FFFFFF")
        else:
            c.fill = fill(COLOURS["alt_row"] if row_num % 2 == 0 else "FFFFFF")
            c.font = font(size=10)
            if "IDOR" in str(val):
                c.fill = fill(COLOURS["high_bg"]); c.font = font(bold=True, color=COLOURS["high_fg"], size=10)
    ws_ep.row_dimensions[row_num].height = 22

# ── Save ──────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"\n[OK] Excel report saved -> {OUTPUT_XLSX}")
print(f"     Sheets: Executive Summary | All Test Results | Findings Only | Endpoint Inventory")
print(f"     Total records : {total}  |  Findings: {len(findings)}  (CRITICAL:{critical_n}  HIGH:{high_n}  MEDIUM:{medium_n})")
